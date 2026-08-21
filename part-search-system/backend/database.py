# -*- coding: utf-8 -*-
"""
van.ea 车辆零件智能查询系统 - 数据库管理模块
"""

import os
import json
import sqlite3
import threading
import re
from datetime import datetime, date
from difflib import SequenceMatcher

import openpyxl

from config import DB_PATH, PART_NUMBER_HEADERS, DB_TYPE

db_lock = threading.Lock()

# json_extract 路径中禁止出现的字符 (防止标识符注入/破坏 SQL 字符串)
_JSON_PATH_ILLEGAL = set('\'"[]\\\n\r\t')


def _json_field(col):
    """
    安全构造字段提取表达式 (双引擎):
      SQLite:      json_extract(data, '$."col"')
      PostgreSQL:  (data->>'col')
    对列名做白名单校验并转义, 防止由列名引发的 SQL 标识符注入。
    非法列名抛出 ValueError (调用方多为管理端配置来源)。
    """
    if not isinstance(col, str) or not col.strip():
        raise ValueError("非法字段名")
    if any(ch in _JSON_PATH_ILLEGAL for ch in col):
        raise ValueError(f"字段名包含非法字符: {col!r}")
    if DB_TYPE == "postgresql":
        # ->>/-> 的右操作数是字符串 key, 列名已过滤单引号/双引号, 直接内插安全
        return f"(data->>'{col}')"
    esc = col.replace('\\', '\\\\').replace('"', '\\"')
    return f'json_extract(data, \'$."{esc}"\')'


def _pg_sql(sql):
    """将 SQLite 风格 SQL 转为 PostgreSQL 语法:
    - 字面 % → %% (psycopg2 参数化转义)
    - 占位符 ? → %s
    """
    return sql.replace('%', '%%').replace('?', '%s')


# ============ NL2SQL: 只读 SQL 安全校验 & 方言转换 ============

# 禁止出现的关键字 (写操作/DDL/危险函数), 用单词边界匹配
_SQL_FORBIDDEN = re.compile(
    r'\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|REPLACE|TRUNCATE|'
    r'ATTACH|DETACH|PRAGMA|GRANT|REVOKE|EXEC|EXECUTE|VACUUM|REINDEX|'
    r'INTO\s+TABLE|COPY|MERGE|CALL)\b',
    re.IGNORECASE,
)
# 只允许出现这些语句前缀
_SQL_ALLOWED_PREFIX = re.compile(
    r'^\s*(SELECT|WITH|EXPLAIN)\b', re.IGNORECASE
)
# 危险的函数/表达式（文件读写、环境变量等）
_SQL_DANGEROUS_FUNCS = re.compile(
    r'\b(load_extension|readfile|writefile|pg_read_file|pg_read_binary_file|'
    r'pg_sleep|pg_terminate_backend|shutdown|sqlite_version)\s*\(',
    re.IGNORECASE,
)
# 允许查询的表白名单
_SQL_ALLOWED_TABLES = {'parts_data', 'uploaded_files', 'unified_columns'}


def _sqlite_to_pg(sql):
    """将 LLM 生成的 SQLite 风格只读 SQL 轻量转换为 PostgreSQL 语法。"""
    s = sql
    # json_extract(data, '$."key"') / json_extract(data, '$.key') -> (data->>'key')
    def _json_repl(m):
        path = m.group(1)
        key = re.sub(r'^\$\.?', '', path).strip().strip('"')
        key = key.replace("'", "''")
        return f"(data->>'{key}')"
    s = re.sub(r"json_extract\s*\(\s*data\s*,\s*'([^']+)'\s*\)",
               _json_repl, s, flags=re.IGNORECASE)
    # CAST(... AS TEXT) 在 PG 中同样有效，无需转换
    # || 连接符通用
    # random() 在 PG 中也是 random()（非 SQLite 的 RANDOM()），统一小写即可
    return s


def validate_readonly_sql(sql):
    """校验 LLM 生成的 SQL 是否为安全的只读查询。
    返回 (ok: bool, error: str|None)。
    """
    if not sql or not sql.strip():
        return False, "empty sql"
    s = sql.strip().rstrip(';').strip()
    if ';' in s:
        # 不允许多语句
        return False, "multiple statements not allowed"
    if not _SQL_ALLOWED_PREFIX.match(s):
        return False, "only SELECT/WITH allowed"
    if _SQL_FORBIDDEN.search(s):
        return False, "forbidden keyword"
    if _SQL_DANGEROUS_FUNCS.search(s):
        return False, "forbidden function"
    if '--' in s or '/*' in s:
        return False, "comments not allowed"
    # 表名白名单：提取 FROM/JOIN 后的标识符
    for m in re.finditer(r'\b(?:FROM|JOIN)\s+([a-zA-Z_][a-zA-Z0-9_]*)', s, re.IGNORECASE):
        tname = m.group(1).lower()
        if tname not in _SQL_ALLOWED_TABLES:
            return False, f"table not allowed: {tname}"
    return True, None


class _PgRow(dict):
    """PostgreSQL 结果行, 兼容 sqlite3.Row 的访问方式 (按列名或下标)"""

    def __getitem__(self, key):
        if isinstance(key, int):
            keys = list(self.keys())
            if 0 <= key < len(keys):
                return dict.__getitem__(self, keys[key])
            raise IndexError("row index out of range")
        return dict.__getitem__(self, key)


class _PgCursor:
    """psycopg2 游标包装, 对齐 sqlite3 cursor 接口 (execute/executescript/fetchall/lastrowid)"""

    def __init__(self, conn, cur):
        self._conn = conn
        self._cur = cur
        self.lastrowid = None
        self.description = None

    def _row(self, row):
        """统一转为 dict 行 (兼容 RealDictRow 与普通 tuple 行)"""
        if row is None:
            return None
        if isinstance(row, _PgRow):
            return row
        if isinstance(row, dict):
            return _PgRow(row)
        cols = [d[0] for d in (self._cur.description or [])]
        return _PgRow(zip(cols, row))

    def execute(self, sql, params=None):
        sql2 = _pg_sql(sql)
        is_insert = sql2.lstrip().upper().startswith('INSERT')
        if is_insert and 'RETURNING' not in sql2.upper():
            # 主键统一为 id (uploaded_files/unified_columns/parts_data)
            sql2 += ' RETURNING id'
        try:
            if params is None:
                self._cur.execute(sql2)
            else:
                self._cur.execute(sql2, list(params) if isinstance(params, tuple) else params)
        finally:
            self.description = self._cur.description
        if is_insert:
            # 取 INSERT ... RETURNING id 的结果作为 lastrowid
            try:
                row = self._cur.fetchone()
                if row is not None:
                    self.lastrowid = row[0]
            except Exception:
                pass
        return self

    def executescript(self, script):
        for stmt in script.split(';'):
            s = stmt.strip()
            if s and not s.startswith('--'):
                self.execute(s)
        return self

    def fetchall(self):
        return [self._row(r) for r in self._cur.fetchall()]

    def fetchone(self):
        return self._row(self._cur.fetchone())

    def close(self):
        try:
            self._cur.close()
        except Exception:
            pass

    @property
    def rowcount(self):
        return self._cur.rowcount


class _PgConn:
    """psycopg2 连接包装, 对齐 sqlite3 连接接口 (execute/executescript/commit/rollback/close)"""

    def __init__(self, conn):
        self._conn = conn

    def cursor(self):
        return _PgCursor(self._conn, self._conn.cursor())

    def execute(self, sql, params=None):
        return self.cursor().execute(sql, params)

    def executescript(self, script):
        return self.cursor().executescript(script)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        try:
            if not self._conn.closed:
                try:
                    self._conn.rollback()  # 丢弃未提交事务, 释放连接
                except Exception:
                    pass
                self._conn.close()
        except Exception:
            pass


def get_db():
    """获取数据库连接 (双引擎): sqlite3 或 PostgreSQL(psycopg2 包装)"""
    if DB_TYPE == "postgresql":
        import psycopg2
        from psycopg2.extras import RealDictCursor
        from config import (
            POSTGRES_HOST, POSTGRES_PORT, POSTGRES_USER,
            POSTGRES_PASSWORD, POSTGRES_DB, POSTGRES_SSLMODE,
        )
        conn = psycopg2.connect(
            host=POSTGRES_HOST, port=POSTGRES_PORT,
            user=POSTGRES_USER, password=POSTGRES_PASSWORD,
            dbname=POSTGRES_DB, sslmode=POSTGRES_SSLMODE,
            cursor_factory=RealDictCursor, connect_timeout=10,
        )
        return _PgConn(conn)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def init_db():
    """初始化数据库 (双引擎 DDL)"""
    with db_lock:
        conn = get_db()
        c = conn.cursor()

        if DB_TYPE == "postgresql":
            c.executescript('''
                CREATE TABLE IF NOT EXISTS uploaded_files (
                    id SERIAL PRIMARY KEY,
                    filename TEXT NOT NULL,
                    original_filename TEXT NOT NULL,
                    upload_date TEXT NOT NULL,
                    sheet_name TEXT,
                    total_rows INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'active',
                    file_type TEXT DEFAULT 'supplementary',
                    stage TEXT
                );

                CREATE TABLE IF NOT EXISTS unified_columns (
                    id SERIAL PRIMARY KEY,
                    english_name TEXT NOT NULL UNIQUE,
                    display_name TEXT NOT NULL,
                    original_names TEXT DEFAULT '[]',
                    is_part_number INTEGER DEFAULT 0,
                    created_date TEXT
                );

                CREATE TABLE IF NOT EXISTS column_mapping (
                    id SERIAL PRIMARY KEY,
                    file_id INTEGER NOT NULL,
                    sheet_name TEXT,
                    original_header TEXT NOT NULL,
                    unified_column_id INTEGER,
                    unified_name TEXT,
                    action TEXT DEFAULT 'mapped',
                    FOREIGN KEY (file_id) REFERENCES uploaded_files(id)
                );

                CREATE TABLE IF NOT EXISTS parts_data (
                    id SERIAL PRIMARY KEY,
                    file_id INTEGER NOT NULL,
                    part_number TEXT,
                    row_number INTEGER,
                    data JSONB,
                    FOREIGN KEY (file_id) REFERENCES uploaded_files(id)
                );

                CREATE INDEX IF NOT EXISTS idx_parts_pn ON parts_data(part_number);
                CREATE INDEX IF NOT EXISTS idx_parts_file ON parts_data(file_id);
                CREATE INDEX IF NOT EXISTS idx_parts_pn_file ON parts_data(part_number, file_id);

                -- JSON 表达式索引 (PostgreSQL: data->>'key')
                CREATE INDEX IF NOT EXISTS idx_data_stage ON parts_data((data->>'Baulos_aggr'));
                CREATE INDEX IF NOT EXISTS idx_data_ec ON parts_data((data->>'BuendelNr'));
                CREATE INDEX IF NOT EXISTS idx_data_fav ON parts_data((data->>'FAV_fav'));
                CREATE INDEX IF NOT EXISTS idx_data_zgs ON parts_data((data->>'ZGS DiaP'));
                CREATE INDEX IF NOT EXISTS idx_data_soma ON parts_data((data->>'SOMA in ZEUS'));
                CREATE INDEX IF NOT EXISTS idx_data_kem ON parts_data((data->>'KEM Number'));
                CREATE INDEX IF NOT EXISTS idx_stage_pn ON parts_data((data->>'Baulos_aggr'), part_number);
            ''')
        else:
            c.executescript('''
                CREATE TABLE IF NOT EXISTS uploaded_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    original_filename TEXT NOT NULL,
                    upload_date TEXT NOT NULL,
                    sheet_name TEXT,
                    total_rows INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'active',
                    file_type TEXT DEFAULT 'supplementary',
                    stage TEXT
                );

                CREATE TABLE IF NOT EXISTS unified_columns (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    english_name TEXT NOT NULL UNIQUE,
                    display_name TEXT NOT NULL,
                    original_names TEXT DEFAULT '[]',
                    is_part_number INTEGER DEFAULT 0,
                    created_date TEXT
                );

                CREATE TABLE IF NOT EXISTS column_mapping (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL,
                    sheet_name TEXT,
                    original_header TEXT NOT NULL,
                    unified_column_id INTEGER,
                    unified_name TEXT,
                    action TEXT DEFAULT 'mapped',
                    FOREIGN KEY (file_id) REFERENCES uploaded_files(id)
                );

                CREATE TABLE IF NOT EXISTS parts_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL,
                    part_number TEXT,
                    row_number INTEGER,
                    data TEXT,
                    FOREIGN KEY (file_id) REFERENCES uploaded_files(id)
                );

                CREATE INDEX IF NOT EXISTS idx_parts_pn ON parts_data(part_number);
                CREATE INDEX IF NOT EXISTS idx_parts_file ON parts_data(file_id);
                CREATE INDEX IF NOT EXISTS idx_parts_pn_file ON parts_data(part_number, file_id);

                -- JSON 表达式索引（加速 Delta 阶段筛选、EC/FAV 查询等）
                -- 注意: SQLite 3.9+ 支持 json_extract 表达式索引
                CREATE INDEX IF NOT EXISTS idx_data_stage ON parts_data(json_extract(data, '$.Baulos_aggr'));
                CREATE INDEX IF NOT EXISTS idx_data_ec ON parts_data(json_extract(data, '$.BuendelNr'));
                CREATE INDEX IF NOT EXISTS idx_data_fav ON parts_data(json_extract(data, '$.FAV_fav'));
                CREATE INDEX IF NOT EXISTS idx_data_zgs ON parts_data(json_extract(data, '$."ZGS DiaP"'));
                CREATE INDEX IF NOT EXISTS idx_data_soma ON parts_data(json_extract(data, '$."SOMA in ZEUS"'));
                CREATE INDEX IF NOT EXISTS idx_data_kem ON parts_data(json_extract(data, '$."KEM Number"'));

                -- 组合索引: 阶段 + PN (加速 Delta 计算)
                CREATE INDEX IF NOT EXISTS idx_stage_pn ON parts_data(json_extract(data, '$.Baulos_aggr'), part_number);
            ''')

        # ===== 数据库迁移：为已有 uploaded_files 表添加新列 =====
        if DB_TYPE == "postgresql":
            existing_cols = [r[0] for r in c.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_name = 'uploaded_files'"
            ).fetchall()]
        else:
            existing_cols = [row[1] for row in c.execute('PRAGMA table_info(uploaded_files)').fetchall()]
        if 'file_type' not in existing_cols:
            c.execute("ALTER TABLE uploaded_files ADD COLUMN file_type TEXT DEFAULT 'supplementary'")
        if 'stage' not in existing_cols:
            c.execute("ALTER TABLE uploaded_files ADD COLUMN stage TEXT")

        conn.commit()
        conn.close()


def serialize_value(val):
    """序列化Excel值为字符串"""
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d %H:%M:%S') if isinstance(val, datetime) else val.strftime('%Y-%m-%d')
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def is_part_number_header(header):
    """检查是否是Part Number列"""
    h_lower = header.lower().strip()
    for pn in PART_NUMBER_HEADERS:
        if pn in h_lower:
            return True
    return False


def header_similarity(h1, h2):
    """计算两个表头的相似度"""
    return SequenceMatcher(None, h1.lower().strip(), h2.lower().strip()).ratio()


def suggest_english_name(original_header):
    """为表头建议英文名称"""
    h = original_header.strip()

    # 已知映射
    known_mappings = {
        'sachnummer': 'Part Number',
        'result.sachnummer': 'Part Number',
        'teilbenennung': 'Part Name',
        'result.teilbenennung': 'Part Name',
        'fehler nr.': 'EC Number',
        'fehler nr': 'EC Number',
        'fav nr.': 'FAV Number',
        'fav nr': 'FAV Number',
        'fav_number': 'FAV Number',
        'fav-beschreibung': 'FAV Description',
        'fav_prio': 'FAV Priority',
        'fav_status': 'FAV Status',
        'fav_description': 'FAV Description',
        'soma in zeus ?': 'SOMA in ZEUS',
        'soma in zeus': 'SOMA in ZEUS',
        'kem': 'KEM',
        'kem nummer': 'KEM Number',
        'kem_nummer': 'KEM Number',
        'kem_status': 'KEM Status',
        'sdr link': 'SDR Link',
        'sdr': 'SDR',
        'br': 'Vehicle Series',
        'result.br': 'Vehicle Series',
        'baureihe': 'Vehicle Series',
        'mg': 'Main Group',
        'result.mg_fav': 'Main Group',
        'prio': 'Priority',
        'status': 'Status',
        'istzustand': 'Current Status',
        'result.istzustand': 'Current Status',
        'sollzustand': 'Future Status',
        'result.sollzustand': 'Future Status',
        'bndverantwortlicher': 'Responsible',
        'result.bndverantwortlicher': 'Responsible',
        'responsible': 'Responsible',
        'processstatus': 'Process Status',
        'current status detail': 'Current Status Detail',
        'future status detail': 'Future Status Detail',
        'pem_aggr': 'PEM Aggregate',
        'part list_zgs': 'Part List ZGS',
        'zgs_diap': 'ZGS DiaP',
        'request number': 'Request Number',
        'request process type': 'Request Process Type',
        'change type': 'Change Type',
        'deviation category': 'Deviation Category',
        'use case': 'Use Case',
        'initiator': 'Initiator',
        'champion / responsible requester': 'Champion / Responsible',
        'planned end date': 'Planned End Date',
        'planned implementation date': 'Planned Implementation Date',
        'created on/at': 'Created At',
        'part code': 'Part Code',
        'part name': 'Part Name',
        'category': 'Category',
        'title': 'Title',
        'count': 'Count',
        'baulos_aggr': 'Build Lot Aggregate',
        'awe_aggr': 'AWE Aggregate',
        'es2_aggr': 'ES2 Aggregate',
        'fav_aggr': 'FAV Aggregate',
        'paket_aggr': 'Package Aggregate',
        'pem_aggr': 'PEM Aggregate',
        'snr_zgs_kem_aggr': 'SNR ZGS KEM Aggregate',
        'fav_fav': 'FAV',
        'favstatuskurz_fav': 'FAV Status Short',
        'favverantwortlicher_fav': 'FAV Responsible',
        'mg_fav': 'Main Group FAV',
        'beschreibung_fav': 'FAV Description',
        'angelegtam_fav': 'FAV Created At',
        'deeplink_fav': 'FAV Deep Link',
        'prioritaet_fnr': 'Priority FNR',
        'solldatumnachhaltigabgestellt_fav': 'FAV Sustainable Fixed Date',
        'soma.fav nr.': 'SOMA FAV Number',
        'soma.fav-beschreibung': 'SOMA FAV Description',
        'soma.fav-link': 'SOMA FAV Link',
        'soma.fav-sd': 'SOMA FAV SD',
        'soma.fav-v.': 'SOMA FAV Version',
        'soma.fav-v. abt.': 'SOMA FAV Version Dept',
        'soma.fehler nr.': 'SOMA Fehler Number',
        'soma.fehler-red.': 'SOMA Fehler Reduction',
        'soma.fehler-red. abt.': 'SOMA Fehler Reduction Dept',
        'soma.mg': 'SOMA Main Group',
        'soma.prio': 'SOMA Priority',
        'aea_acm': 'AEA ACM',
        'zgs_acm': 'ZGS ACM',
        'zgs_diap': 'ZGS DiaP',
        'zgs_kem': 'ZGS KEM',
        'zgs_diap_max_vit': 'ZGS DiaP Max Vit',
        'zgs_diap_gesteuert_baulos': 'ZGS DiaP Controlled Build Lot',
        'buendelnr': 'Bundle Number',
        'schritt': 'Step',
        'werk': 'Plant',
        'd': 'Date',
        'erstelldatum_bnd_acm': 'ACM Created Date',
        'beschlussdatum_acm': 'ACM Decision Date',
        'statusproduktivekemfreigabefuz': 'Productive KEM Release Status',
        'stichwortbenennung': 'Keyword Designation',
        'sachnummerhistorieallepv': 'Part Number History All PV',
        'count of fav_fav': 'Count of FAV',
        'responsible requester is qe': 'Responsible Requester is QE',
        'row labels': 'Row Labels',
        'soMA.FAV-SD': 'SOMA FAV SD',
    }

    h_lower = h.lower()
    if h_lower in known_mappings:
        return known_mappings[h_lower]

    # 如果已经是英文，直接返回
    if re.match(r'^[A-Za-z][A-Za-z0-9\s\-\._/]*$', h):
        return h

    # 通用处理：保留原始
    return h


class DatabaseManager:
    """数据库管理器"""

    def __init__(self):
        init_db()

    # ===== 文件管理 =====

    def list_files(self):
        """列出所有上传的文件"""
        conn = get_db()
        rows = conn.execute(
            'SELECT * FROM uploaded_files ORDER BY upload_date DESC'
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def delete_file(self, file_id):
        """删除文件及其数据"""
        with db_lock:
            conn = get_db()
            conn.execute('DELETE FROM parts_data WHERE file_id = ?', (file_id,))
            conn.execute('DELETE FROM column_mapping WHERE file_id = ?', (file_id,))
            conn.execute('DELETE FROM uploaded_files WHERE id = ?', (file_id,))
            conn.commit()
            conn.close()

    # ===== Excel上传预处理 =====

    def analyze_excel(self, filepath, original_filename):
        """
        分析Excel文件，返回所有工作表的表头信息
        不导入数据，只做预处理
        """
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheets_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            # 使用 iter_rows 获取第一行（read_only 模式兼容）
            first_row = next(ws.iter_rows(min_row=1, max_row=1), [])
            for cell in first_row:
                headers.append(serialize_value(cell.value))

            # 跳过空表
            if not headers or all(h == '' for h in headers):
                continue

            # 获取行数估算（对超大表只取前几行判断是否非空）
            row_count = 0
            for _ in ws.iter_rows(min_row=2, max_row=10):
                row_count += 1
            # 使用 max_row 估算，限制最大值避免超大表卡死
            try:
                max_row = ws.max_row
                if max_row and max_row > 0:
                    total_rows = min(max_row - 1, 999999)
                else:
                    total_rows = row_count
            except Exception:
                total_rows = row_count

            sheets_info.append({
                'sheet_name': sheet_name,
                'headers': headers,
                'estimated_rows': total_rows
            })

        wb.close()
        return sheets_info

    def get_existing_columns(self):
        """获取现有的统一列"""
        conn = get_db()
        rows = conn.execute('SELECT * FROM unified_columns ORDER BY id').fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def create_column_mapping(self, file_headers, sheet_name):
        """
        为Excel表头创建列映射
        返回: {matched: [...], unmatched: [...]}
        """
        existing_cols = self.get_existing_columns()
        existing_names = {col['english_name'].lower(): col for col in existing_cols}

        matched = []
        unmatched = []

        for header in file_headers:
            if not header or header.strip() == '':
                continue

            h_lower = header.lower().strip()
            eng_name = suggest_english_name(header)
            eng_lower = eng_name.lower()

            # 1. 精确匹配现有列
            if eng_lower in existing_names:
                col = existing_names[eng_lower]
                matched.append({
                    'original_header': header,
                    'unified_column_id': col['id'],
                    'unified_name': col['english_name'],
                    'display_name': col['display_name'],
                    'similarity': 1.0,
                    'is_part_number': col['is_part_number']
                })
                continue

            # 2. 模糊匹配（相似度 > 0.7）
            best_match = None
            best_score = 0
            for col in existing_cols:
                score = header_similarity(header, col['english_name'])
                if score > best_score:
                    best_score = score
                    best_match = col

            if best_match and best_score >= 0.7:
                matched.append({
                    'original_header': header,
                    'unified_column_id': best_match['id'],
                    'unified_name': best_match['english_name'],
                    'display_name': best_match['display_name'],
                    'similarity': best_score,
                    'is_part_number': best_match['is_part_number']
                })
            else:
                # 无法匹配，需要管理员选择
                is_pn = is_part_number_header(header)
                unmatched.append({
                    'original_header': header,
                    'suggested_english': eng_name,
                    'is_part_number': is_pn,
                    'best_match_name': best_match['english_name'] if best_match else None,
                    'best_match_score': best_score,
                    'display_name': f"{eng_name} / {header}" if eng_name.lower() != h_lower else header
                })

        return {
            'matched': matched,
            'unmatched': unmatched,
            'sheet_name': sheet_name
        }

    def import_excel_data(self, filepath, original_filename, sheet_mappings, admin_english_names=None,
                          file_type='supplementary', stage=None):
        """
        导入Excel数据到数据库
        sheet_mappings: {sheet_name: [{original_header, unified_name, action}, ...]}
        admin_english_names: {original_header: english_name} 管理员指定的英文名
        file_type: 'BOM' 或 'supplementary'，标记文件类型
        stage: 'pre-TO' / 'TO1' / 'TO2'，BOM文件所属阶段（supplementary文件为 None）
        """
        admin_english_names = admin_english_names or {}
        results = []

        wb = openpyxl.load_workbook(filepath, data_only=True)

        for sheet_name, mappings in sheet_mappings.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [serialize_value(cell.value) for cell in ws[1]]

            # 注册文件
            with db_lock:
                conn = get_db()
                cursor = conn.execute(
                    'INSERT INTO uploaded_files (filename, original_filename, upload_date, sheet_name, total_rows, status, file_type, stage) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                    (os.path.basename(filepath), original_filename, datetime.now().isoformat(), sheet_name, ws.max_row - 1, 'active', file_type, stage)
                )
                file_id = cursor.lastrowid

                # 处理映射：创建统一列 + 列映射
                header_to_unified = {}
                for m in mappings:
                    orig = m['original_header']
                    unified_name = m.get('unified_name') or m.get('suggested_english') or orig
                    action = m.get('action', 'mapped')

                    if action == 'skip':
                        header_to_unified[orig] = None
                        continue

                    # 检查统一列是否已存在
                    col_row = conn.execute(
                        'SELECT * FROM unified_columns WHERE english_name = ?', (unified_name,)
                    ).fetchone()

                    if col_row:
                        col_id = col_row['id']
                        # 更新original_names
                        orig_names = json.loads(col_row['original_names'] or '[]')
                        if orig not in orig_names:
                            orig_names.append(orig)
                        display = col_row['display_name']
                        # 如果管理员指定了英文名，更新display
                        if orig in admin_english_names:
                            eng = admin_english_names[orig]
                            display = f"{eng} / {orig}" if eng.lower() != orig.lower() else orig
                            conn.execute(
                                'UPDATE unified_columns SET display_name = ?, original_names = ? WHERE id = ?',
                                (display, json.dumps(orig_names), col_id)
                            )
                        else:
                            conn.execute(
                                'UPDATE unified_columns SET original_names = ? WHERE id = ?',
                                (json.dumps(orig_names), col_id)
                            )
                    else:
                        # 创建新列
                        is_pn = is_part_number_header(orig) or is_part_number_header(unified_name)
                        eng_name = admin_english_names.get(orig, unified_name)
                        display = f"{eng_name} / {orig}" if eng_name.lower() != orig.lower() else orig
                        cursor2 = conn.execute(
                            'INSERT INTO unified_columns (english_name, display_name, original_names, is_part_number, created_date) VALUES (?, ?, ?, ?, ?)',
                            (unified_name, display, json.dumps([orig]), is_pn, datetime.now().isoformat())
                        )
                        col_id = cursor2.lastrowid

                    # 创建列映射记录
                    conn.execute(
                        'INSERT INTO column_mapping (file_id, sheet_name, original_header, unified_column_id, unified_name, action) VALUES (?, ?, ?, ?, ?, ?)',
                        (file_id, sheet_name, orig, col_id, unified_name, action)
                    )

                    header_to_unified[orig] = unified_name

                # 导入数据行
                imported = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_data = {}
                    part_number = ''
                    for ci, header in enumerate(headers):
                        if header not in header_to_unified:
                            continue
                        unified_name = header_to_unified[header]
                        if unified_name is None:
                            continue
                        val = serialize_value(row[ci]) if ci < len(row) else ''
                        row_data[unified_name] = val
                        if is_part_number_header(header) or is_part_number_header(unified_name):
                            if not part_number:
                                part_number = val

                    # 跳过完全空的行
                    if not any(v for v in row_data.values()):
                        continue

                    conn.execute(
                        'INSERT INTO parts_data (file_id, part_number, row_number, data) VALUES (?, ?, ?, ?)',
                        (file_id, part_number, row_idx, json.dumps(row_data, ensure_ascii=False))
                    )
                    imported += 1

                conn.execute('UPDATE uploaded_files SET total_rows = ? WHERE id = ?', (imported, file_id))
                conn.commit()
                conn.close()

            results.append({
                'sheet_name': sheet_name,
                'file_id': file_id,
                'imported_rows': imported
            })

        wb.close()

        # 导入后清理：检测并删除全空列
        removed_cols = self._remove_empty_columns()
        if removed_cols:
            print(f"[DB] 导入后清理：已删除 {len(removed_cols)} 个全空列: {removed_cols}")

        return results

    def _remove_empty_columns(self):
        """
        检测并删除全空列（在所有记录中均无数据的统一列）
        返回被删除的列名列表
        """
        with db_lock:
            conn = get_db()
            columns = conn.execute('SELECT id, english_name FROM unified_columns').fetchall()
            removed = []

            for col in columns:
                col_name = col['english_name']
                # 跳过Part Number列（关键键不应删除）
                if is_part_number_header(col_name):
                    continue

                # 检查该列在所有记录中是否有非空数据
                cnt = conn.execute(
                    f"SELECT COUNT(*) as c FROM parts_data "
                    f"WHERE {_json_field(col_name)} IS NOT NULL "
                    f"AND {_json_field(col_name)} != ''"
                ).fetchone()['c']

                if cnt == 0:
                    # 删除全空列
                    conn.execute('DELETE FROM unified_columns WHERE id = ?', (col['id'],))
                    conn.execute('DELETE FROM column_mapping WHERE unified_column_id = ?', (col['id'],))
                    removed.append(col_name)

            conn.commit()
            conn.close()
            return removed

    # ===== 查询 =====

    def search_by_part_number(self, part_number, exact=False):
        """按零件号搜索"""
        conn = get_db()
        if exact:
            rows = conn.execute(
                'SELECT * FROM parts_data WHERE part_number = ? ORDER BY file_id',
                (part_number,)
            ).fetchall()
        else:
            rows = conn.execute(
                'SELECT * FROM parts_data WHERE part_number LIKE ? ORDER BY file_id',
                (f'%{part_number}%',)
            ).fetchall()

        results = []
        for row in rows:
            data = json.loads(row['data'])
            data['_file_id'] = row['file_id']
            data['_row_number'] = row['row_number']
            data['_record_id'] = row['id']
            results.append(data)

        # 获取文件名映射
        file_ids = set(r['_file_id'] for r in results)
        file_names = {}
        if file_ids:
            placeholders = ','.join('?' * len(file_ids))
            file_rows = conn.execute(
                f'SELECT id, original_filename, sheet_name FROM uploaded_files WHERE id IN ({placeholders})',
                list(file_ids)
            ).fetchall()
            file_names = {r['id']: {'filename': r['original_filename'], 'sheet': r['sheet_name']} for r in file_rows}

        conn.close()

        for r in results:
            fid = r['_file_id']
            r['_source_file'] = file_names.get(fid, {}).get('filename', '')
            r['_source_sheet'] = file_names.get(fid, {}).get('sheet', '')

        return results

    def search_by_field(self, field_name, value, exact=False):
        """按任意字段搜索（使用JSON查询, 兼容双引擎）"""
        conn = get_db()

        # 字段名经 _json_field 校验, 防止标识符注入
        try:
            field_expr = _json_field(field_name)
        except ValueError:
            conn.close()
            return []

        if exact:
            query = f"SELECT * FROM parts_data WHERE {field_expr} = ?"
            rows = conn.execute(query, (value,)).fetchall()
        else:
            query = f"SELECT * FROM parts_data WHERE CAST({field_expr} AS TEXT) LIKE ?"
            rows = conn.execute(query, (f'%{value}%',)).fetchall()

        results = []
        for row in rows:
            data = json.loads(row['data'])
            data['_file_id'] = row['file_id']
            data['_row_number'] = row['row_number']
            data['_record_id'] = row['id']
            results.append(data)

        file_ids = set(r['_file_id'] for r in results)
        file_names = {}
        if file_ids:
            placeholders = ','.join('?' * len(file_ids))
            file_rows = conn.execute(
                f'SELECT id, original_filename, sheet_name FROM uploaded_files WHERE id IN ({placeholders})',
                list(file_ids)
            ).fetchall()
            file_names = {r['id']: {'filename': r['original_filename'], 'sheet': r['sheet_name']} for r in file_rows}

        conn.close()
        for r in results:
            fid = r['_file_id']
            r['_source_file'] = file_names.get(fid, {}).get('filename', '')
            r['_source_sheet'] = file_names.get(fid, {}).get('sheet', '')

        return results

    def get_all_columns(self):
        """获取所有统一列"""
        conn = get_db()
        rows = conn.execute('SELECT * FROM unified_columns ORDER BY is_part_number DESC, id').fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def get_stats(self):
        """获取数据库统计"""
        conn = get_db()
        file_count = conn.execute('SELECT COUNT(*) as c FROM uploaded_files').fetchone()['c']
        record_count = conn.execute('SELECT COUNT(*) as c FROM parts_data').fetchone()['c']
        col_count = conn.execute('SELECT COUNT(*) as c FROM unified_columns').fetchone()['c']
        pn_count = conn.execute("SELECT COUNT(DISTINCT part_number) as c FROM parts_data WHERE part_number != ''").fetchone()['c']

        files = conn.execute('SELECT id, original_filename, sheet_name, total_rows, upload_date FROM uploaded_files ORDER BY upload_date DESC').fetchall()
        conn.close()

        return {
            'total_files': file_count,
            'total_records': record_count,
            'total_columns': col_count,
            'unique_part_numbers': pn_count,
            'files': [dict(f) for f in files]
        }

    def get_all_part_numbers(self):
        """获取所有零件号"""
        conn = get_db()
        rows = conn.execute(
            "SELECT DISTINCT part_number FROM parts_data WHERE part_number != '' AND part_number IS NOT NULL ORDER BY part_number"
        ).fetchall()
        conn.close()
        return [r['part_number'] for r in rows]

    def get_search_hint_samples(self, limit_per_field=3):
        """获取可搜索字段的样例值，用于生成搜索提示词。

        返回: dict，键为业务字段英文名，值为 {"label": 显示名, "samples": [样例值...]}
        """
        # 可搜索的核心业务字段（英文标识名 -> 显示名）
        searchable_fields = {
            'part_number': 'Part Number',
            'BuendelNr': 'EC',
            'KEM': 'KEM',
            'FAV_fav': 'FAV',
            'ZGS DiaP': 'ZGS',
            'SOMA in ZEUS': 'SOMA',
            'Baulos_aggr': 'Stage',
            'bndverantwortlicher': 'Responsible',
            'status': 'Status',
            'teilbenennung': 'Part Name',
        }
        conn = get_db()
        result = {}
        try:
            for field_en, label in searchable_fields.items():
                try:
                    if field_en == 'part_number':
                        rows = conn.execute(
                            "SELECT DISTINCT part_number AS v FROM parts_data "
                            "WHERE part_number != '' AND part_number IS NOT NULL "
                            "ORDER BY RANDOM() LIMIT ?",
                            (limit_per_field,)
                        ).fetchall()
                    else:
                        field_expr = _json_field(field_en)
                        sql = (f"SELECT DISTINCT {field_expr} AS v FROM parts_data "
                               f"WHERE {field_expr} IS NOT NULL "
                               f"AND CAST({field_expr} AS TEXT) != '' "
                               f"ORDER BY RANDOM() LIMIT ?")
                        rows = conn.execute(sql, (limit_per_field,)).fetchall()
                    samples = [str(r['v']).strip() for r in rows
                               if r['v'] is not None and str(r['v']).strip()]
                    if samples:
                        result[field_en] = {'label': label, 'samples': samples}
                except Exception:
                    continue
        finally:
            conn.close()
        return result

    def get_schema_description(self):
        """返回供 LLM NL2SQL 使用的数据库 schema 文本描述。"""
        conn = get_db()
        try:
            # 从实际数据中抽样 JSON key 及其样例值
            sample_row = conn.execute(
                "SELECT data FROM parts_data WHERE data IS NOT NULL LIMIT 1"
            ).fetchone()
            sample_keys = []
            if sample_row and sample_row['data']:
                try:
                    sample_keys = list(json.loads(sample_row['data']).keys())
                except Exception:
                    sample_keys = []

            # 常用关键字段的中文含义（辅助 LLM 理解）
            key_hints = {
                'BuendelNr': 'EC编号/错误号 (EC / Fehler Nr.)',
                'KEM': 'KEM编号',
                'FAV_fav': 'FAV编号',
                'ZGS DiaP': 'ZGS版本号',
                'SOMA in ZEUS': 'SOMA状态(ja/nein)',
                'Baulos_aggr': '阶段/批次 (Stage/Baulos)',
                'bndverantwortlicher': '负责人',
                'status': '状态',
                'teilbenennung': '零件名称',
                'Teilbenennung': '零件名称',
                'Sachnummer': '零件号（冗余在 part_number 列）',
                'MG': '主组编号',
                'BR': '车型系列',
            }

            lines = [
                "表: parts_data (零件数据表)",
                "结构化列:",
                "  id          INTEGER  主键",
                "  part_number TEXT     零件号 (Sachnummer)",
                "  file_id     INTEGER  来源文件ID",
                "  row_number  INTEGER  行号",
                "  data        JSON/TEXT  其余业务字段，使用 json_extract 访问",
                "",
                "JSON data 字段访问方式: json_extract(data, '$.\"字段名\"')",
                "示例: json_extract(data, '$.\"BuendelNr\"')",
                "模糊匹配: CAST(json_extract(data, '$.\"BuendelNr\"') AS TEXT) LIKE '%值%'",
                "非空判断: json_extract(data, '$.\"BuendelNr\"') IS NOT NULL",
                "",
                "data 中常见的字段:",
            ]
            shown = set()
            for k in sample_keys[:40]:
                hint = key_hints.get(k, '')
                lines.append(f"  - {k}" + (f"  -- {hint}" if hint else ''))
                shown.add(k)
            # 补充未出现在 sample 但已知的重要字段
            for k, hint in key_hints.items():
                if k not in shown:
                    lines.append(f"  - {k}  -- {hint}")

            lines += [
                "",
                "其他可用表: uploaded_files(id, filename, original_filename, sheet_name, total_rows, upload_date, stage), unified_columns(english_name, display_name)",
                "规则:",
                "  1. 只生成 SELECT 语句，不要写操作",
                "  2. 业务数据查询始终加 LIMIT，最大 100 行",
                "  3. 计数/统计可用 COUNT(*), 不需要 LIMIT",
                "  4. 字段名可能含空格或特殊字符，JSON 路径中用双引号包裹",
            ]
            return '\n'.join(lines)
        finally:
            conn.close()

    def execute_readonly_sql(self, sql):
        """执行 LLM 生成的只读 SQL。
        返回 (results: list[dict], columns: list[str], error: str|None)。
        - 自动进行安全校验
        - 自动为无 LIMIT 的 SELECT 添加 LIMIT 100
        - 聚合查询（含 COUNT/SUM/AVG/MIN/MAX/分组）不加 LIMIT
        - 自动转换 SQLite JSON 语法为 PostgreSQL 语法
        """
        ok, err = validate_readonly_sql(sql)
        if not ok:
            return [], [], err

        s = sql.strip().rstrip(';').strip()

        # 检测是否为聚合查询（不强制加 LIMIT）
        is_aggregate = bool(
            re.search(r'\b(COUNT|SUM|AVG|MIN|MAX|GROUP\s+BY)\b', s, re.IGNORECASE)
        )
        if not is_aggregate and not re.search(r'\bLIMIT\b', s, re.IGNORECASE):
            s += ' LIMIT 100'

        if DB_TYPE == 'postgresql':
            s = _sqlite_to_pg(s)

        conn = get_db()
        try:
            cur = conn.execute(s)
            rows = cur.fetchall()
            cols = [d[0] for d in (cur.description or [])]
            # 对于 parts_data 行，展开 JSON 为业务字段（和 search_by_field 保持一致）
            results = []
            is_parts_rows = ('id' in cols and 'part_number' in cols and 'data' in cols)
            if is_parts_rows:
                file_ids = set()
                parsed = []
                for row in rows:
                    rd = dict(row)
                    data = {}
                    try:
                        data = json.loads(rd.get('data') or '{}')
                    except Exception:
                        data = {}
                    data['_file_id'] = rd.get('file_id')
                    data['_row_number'] = rd.get('row_number')
                    data['_record_id'] = rd.get('id')
                    data.setdefault('part_number', rd.get('part_number', ''))
                    parsed.append(data)
                    if data['_file_id'] is not None:
                        file_ids.add(data['_file_id'])
                file_names = {}
                if file_ids:
                    placeholders = ','.join('?' * len(file_ids))
                    fsql = (f'SELECT id, original_filename, sheet_name '
                            f'FROM uploaded_files WHERE id IN ({placeholders})')
                    if DB_TYPE == 'postgresql':
                        fsql = _pg_sql(fsql)
                    frows = conn.execute(fsql, list(file_ids)).fetchall()
                    file_names = {r['id']: {'filename': r['original_filename'],
                                            'sheet': r['sheet_name']} for r in frows}
                for r in parsed:
                    fid = r.get('_file_id')
                    r['_source_file'] = file_names.get(fid, {}).get('filename', '')
                    r['_source_sheet'] = file_names.get(fid, {}).get('sheet', '')
                    results.append(r)
                # 更新 columns 为展开后的业务字段（取首条 keys）
                if results:
                    cols = [k for k in results[0].keys() if not k.startswith('_source')]
            else:
                for row in rows:
                    results.append({k: row[k] for k in cols})
            return results, cols, None
        except Exception as e:
            return [], [], str(e)
        finally:
            conn.close()

    def update_cell(self, record_id, field_name, value):
        """更新单个单元格"""
        with db_lock:
            conn = get_db()
            row = conn.execute('SELECT data FROM parts_data WHERE id = ?', (record_id,)).fetchone()
            if not row:
                conn.close()
                return False

            data = json.loads(row['data'])
            data[field_name] = value

            # 如果更新的是Part Number，同步更新索引列
            if is_part_number_header(field_name):
                conn.execute('UPDATE parts_data SET data = ?, part_number = ? WHERE id = ?',
                             (json.dumps(data, ensure_ascii=False), value, record_id))
            else:
                conn.execute('UPDATE parts_data SET data = ? WHERE id = ?',
                             (json.dumps(data, ensure_ascii=False), record_id))

            conn.commit()
            conn.close()
            return True

    def update_column_display(self, col_id, english_name, display_name):
        """更新列的显示名称"""
        with db_lock:
            conn = get_db()
            conn.execute(
                'UPDATE unified_columns SET english_name = ?, display_name = ? WHERE id = ?',
                (english_name, display_name, col_id)
            )
            conn.commit()
            conn.close()

    # ===== 仪表盘统计 =====

    def _find_best_column(self, conn, col_names, keywords):
        """在多个候选列中找到数据最多的列"""
        candidates = [c for c in col_names if any(kw in c.lower() for kw in keywords)]
        if not candidates:
            return None
        best_col = None
        best_count = 0
        for col in candidates:
            cnt = conn.execute(
                f"SELECT COUNT(*) as c FROM parts_data WHERE {_json_field(col)} IS NOT NULL AND {_json_field(col)} != ''"
            ).fetchone()['c']
            if cnt > best_count:
                best_count = cnt
                best_col = col
        return best_col if best_col else candidates[0]

    def get_dashboard_stats(self):
        """获取仪表盘详细统计数据"""
        conn = get_db()
        columns = conn.execute('SELECT english_name, display_name FROM unified_columns').fetchall()
        col_names = [c['english_name'] for c in columns]
        col_display = {c['english_name']: c['display_name'] for c in columns}

        # 基本统计
        total_records = conn.execute('SELECT COUNT(*) as c FROM parts_data').fetchone()['c']
        unique_pn = conn.execute(
            "SELECT COUNT(DISTINCT part_number) as c FROM parts_data WHERE part_number != '' AND part_number IS NOT NULL"
        ).fetchone()['c']

        # 使用数据最多的列进行统计
        ec_col = self._find_best_column(conn, col_names, ['ec', 'fehler'])
        fav_col = self._find_best_column(conn, col_names, ['fav', 'zeus'])
        soma_col = self._find_best_column(conn, col_names, ['soma'])
        baulos_col = self._find_best_column(conn, col_names, ['baulos'])

        stats = {
            'total_records': total_records,
            'unique_part_numbers': unique_pn,
            'total_files': conn.execute('SELECT COUNT(*) as c FROM uploaded_files').fetchone()['c'],
            'total_columns': len(col_names),
            'ec_count': 0,
            'ec_pn_count': 0,
            'fav_count': 0,
            'soma_count': 0,
            'soma_yes': 0,
            'soma_no': 0,
            'field_distribution': {},
            'source_distribution': [],
            'ec_col_name': ec_col,
            'fav_col_name': fav_col,
            'soma_col_name': soma_col,
            'baulos_col_name': baulos_col,
            'col_display': col_display,
            'baulos_stages': [],
        }

        # EC统计 - 使用数据最多的EC列
        if ec_col:
            ec_query = f"SELECT COUNT(DISTINCT {_json_field(ec_col)}) as c FROM parts_data WHERE {_json_field(ec_col)} IS NOT NULL AND {_json_field(ec_col)} != ''"
            stats['ec_count'] = conn.execute(ec_query).fetchone()['c']

            ec_pn_query = f"SELECT COUNT(DISTINCT part_number) as c FROM parts_data WHERE part_number != '' AND {_json_field(ec_col)} IS NOT NULL AND {_json_field(ec_col)} != ''"
            stats['ec_pn_count'] = conn.execute(ec_pn_query).fetchone()['c']

        # FAV/ZEUS统计 - 使用数据最多的FAV列
        if fav_col:
            fav_query = f"SELECT COUNT(DISTINCT {_json_field(fav_col)}) as c FROM parts_data WHERE {_json_field(fav_col)} IS NOT NULL AND {_json_field(fav_col)} != ''"
            stats['fav_count'] = conn.execute(fav_query).fetchone()['c']

        # SOMA统计
        if soma_col:
            soma_query = f"SELECT {_json_field(soma_col)} as val, COUNT(*) as c FROM parts_data GROUP BY {_json_field(soma_col)}"
            soma_rows = conn.execute(soma_query).fetchall()
            stats['soma_count'] = sum(r['c'] for r in soma_rows if r['val'] and r['val'].strip())
            for r in soma_rows:
                val = (r['val'] or '').strip().lower()
                if val in ['ja', 'yes', 'true', '1', 'x']:
                    stats['soma_yes'] += r['c']
                elif val in ['nein', 'no', 'false', '0', '']:
                    stats['soma_no'] += r['c']
            stats['soma_distribution'] = [
                {'value': r['val'] or '(空)', 'count': r['c']}
                for r in soma_rows if r['c'] > 0
            ]

        # Baulos阶段统计 (PRO1=TO1, PRO2=TO2)
        if baulos_col:
            baulos_rows = conn.execute(
                f"SELECT {_json_field(baulos_col)} as val, COUNT(*) as total, "
                f"COUNT(DISTINCT part_number) as pn_count FROM parts_data "
                f"GROUP BY {_json_field(baulos_col)}"
            ).fetchall()

            # 按PRO1/PRO2阶段聚合
            pro1_total = 0
            pro1_pn = 0
            pro2_total = 0
            pro2_pn = 0

            for r in baulos_rows:
                val = (r['val'] or '').upper()
                if 'PRO1' in val:
                    pro1_total += r['total']
                    pro1_pn += r['pn_count']
                if 'PRO2' in val:
                    pro2_total += r['total']
                    pro2_pn += r['pn_count']

            # EC按阶段统计 + 解决情况追踪
            pro1_ec_count = 0
            pro2_ec_count = 0
            ec_resolved = 0
            ec_remaining = 0

            if ec_col:
                # TO1阶段的所有不同EC
                pro1_ec_rows = conn.execute(
                    f"SELECT DISTINCT {_json_field(ec_col)} as ec "
                    f"FROM parts_data "
                    f"WHERE UPPER({_json_field(baulos_col)}) LIKE '%PRO1%' "
                    f"AND {_json_field(ec_col)} IS NOT NULL "
                    f"AND {_json_field(ec_col)} != ''"
                ).fetchall()
                pro1_ec_set = {r['ec'] for r in pro1_ec_rows}
                pro1_ec_count = len(pro1_ec_set)

                # TO2阶段的所有不同EC
                pro2_ec_rows = conn.execute(
                    f"SELECT DISTINCT {_json_field(ec_col)} as ec "
                    f"FROM parts_data "
                    f"WHERE UPPER({_json_field(baulos_col)}) LIKE '%PRO2%' "
                    f"AND {_json_field(ec_col)} IS NOT NULL "
                    f"AND {_json_field(ec_col)} != ''"
                ).fetchall()
                pro2_ec_set = {r['ec'] for r in pro2_ec_rows}
                pro2_ec_count = len(pro2_ec_set)

                # EC解决情况：
                # 已解决 = 在TO1中存在但不在TO2中的EC
                ec_resolved = len(pro1_ec_set - pro2_ec_set)
                # 仍未解决 = 同时存在于TO1和TO2中的EC
                ec_remaining = len(pro1_ec_set & pro2_ec_set)

            stats['baulos_stages'] = [
                {'stage': 'PRO1 (TO1)', 'total': pro1_total, 'pn_count': pro1_pn, 'ec_count': pro1_ec_count},
                {'stage': 'PRO2 (TO2)', 'total': pro2_total, 'pn_count': pro2_pn, 'ec_count': pro2_ec_count},
            ]
            stats['ec_resolution'] = {
                'to1_ec_count': pro1_ec_count,
                'to2_ec_count': pro2_ec_count,
                'resolved': ec_resolved,
                'remaining': ec_remaining,
            }

        # 来源文件分布
        file_dist = conn.execute(
            'SELECT uf.original_filename, uf.sheet_name, COUNT(pd.id) as cnt '
            'FROM uploaded_files uf LEFT JOIN parts_data pd ON uf.id = pd.file_id '
            'GROUP BY uf.id ORDER BY cnt DESC'
        ).fetchall()
        stats['source_distribution'] = [
            {'label': f"{r['original_filename']} / {r['sheet_name']}", 'count': r['cnt']}
            for r in file_dist
        ]

        # 字段覆盖率（前15个字段）
        for col_name in col_names[:15]:
            non_empty = conn.execute(
                f"SELECT COUNT(*) as c FROM parts_data WHERE {_json_field(col_name)} IS NOT NULL AND {_json_field(col_name)} != ''"
            ).fetchone()['c']
            stats['field_distribution'][col_name] = {
                'filled': non_empty,
                'total': total_records,
                'rate': round(non_empty / total_records * 100, 1) if total_records > 0 else 0
            }

        conn.close()
        return stats

    # ===== 仪表盘下钻查询 =====

    def get_drilldown_data(self, dimension, value, page=1, page_size=20):
        """
        仪表盘下钻查询：按维度和值获取详细记录
        dimension: 'soma' | 'ec' | 'fav' | 'source' | 'field' | 'all'
        value: 具体的值（如 'ja', 'nein', 'has_ec', 'no_ec', source_label, field_name）
        返回: {records, total, page, page_size, dimension, value}
        """
        conn = get_db()
        columns = conn.execute('SELECT english_name, display_name FROM unified_columns').fetchall()
        col_names = [c['english_name'] for c in columns]
        col_display = {c['english_name']: c['display_name'] for c in columns}

        where_clause = "1=1"
        params = []
        title_desc = ""

        if dimension == 'soma':
            soma_cols = [c for c in col_names if 'soma' in c.lower()]
            if not soma_cols:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'SOMA column not found'}
            soma_col = soma_cols[0]
            if value.lower() in ['ja', 'yes', '是']:
                where_clause = f"LOWER({_json_field(soma_col)}) IN ('ja', 'yes', 'true', '1', 'x')"
                title_desc = f"SOMA = Ja"
            else:
                where_clause = f"({_json_field(soma_col)} IS NULL OR {_json_field(soma_col)} = '' OR LOWER({_json_field(soma_col)}) IN ('nein', 'no', 'false', '0'))"
                title_desc = f"SOMA = Nein/空"

        elif dimension == 'ec':
            ec_cols = [c for c in col_names if 'ec' in c.lower()]
            if not ec_cols:
                ec_cols = [c for c in col_names if 'fehler' in c.lower()]
            if not ec_cols:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'EC column not found'}
            ec_col = ec_cols[0]
            if value == 'has_ec':
                where_clause = f"{_json_field(ec_col)} IS NOT NULL AND {_json_field(ec_col)} != ''"
                title_desc = f"含EC的零件 (EC列: {col_display.get(ec_col, ec_col)})"
            else:
                where_clause = f"({_json_field(ec_col)} IS NULL OR {_json_field(ec_col)} = '')"
                title_desc = f"不含EC的零件"

        elif dimension == 'fav':
            fav_cols = [c for c in col_names if 'fav' in c.lower() or 'zeus' in c.lower()]
            if not fav_cols:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'FAV column not found'}
            fav_col = fav_cols[0]
            if value == 'has_fav':
                where_clause = f"{_json_field(fav_col)} IS NOT NULL AND {_json_field(fav_col)} != ''"
                title_desc = f"含FAV/ZEUS的零件 (列: {col_display.get(fav_col, fav_col)})"
            else:
                where_clause = f"({_json_field(fav_col)} IS NULL OR {_json_field(fav_col)} = '')"
                title_desc = f"不含FAV/ZEUS的零件"

        elif dimension == 'source':
            # value is the source label: "filename / sheet"
            files = conn.execute('SELECT id, original_filename, sheet_name FROM uploaded_files').fetchall()
            matched_file_id = None
            for f in files:
                label = f"{f['original_filename']} / {f['sheet_name']}"
                if label == value:
                    matched_file_id = f['id']
                    break
            if matched_file_id is None:
                # Try partial match
                for f in files:
                    label = f"{f['original_filename']} / {f['sheet_name']}"
                    if value in label or label in value:
                        matched_file_id = f['id']
                        break
            if matched_file_id is None:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'Source not found'}
            where_clause = f"file_id = ?"
            params = [matched_file_id]
            title_desc = f"来源: {value}"

        elif dimension == 'field':
            # value is the field name; show records where this field is filled
            if value not in col_names:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'Field not found'}
            where_clause = f"{_json_field(value)} IS NOT NULL AND {_json_field(value)} != ''"
            title_desc = f"字段已填充: {col_display.get(value, value)}"

        elif dimension == 'baulos':
            # value is 'pro1' or 'pro2'
            baulos_cols = [c for c in col_names if 'baulos' in c.lower()]
            if not baulos_cols:
                conn.close()
                return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                        'dimension': dimension, 'value': value, 'error': 'Baulos column not found'}
            baulos_col = baulos_cols[0]
            val_upper = value.upper()
            where_clause = f"UPPER({_json_field(baulos_col)}) LIKE ?"
            params = [f'%{val_upper}%']
            title_desc = f"Baulos阶段: {val_upper}"

        elif dimension == 'all':
            title_desc = "所有记录"
        else:
            conn.close()
            return {'records': [], 'total': 0, 'page': page, 'page_size': page_size,
                    'dimension': dimension, 'value': value, 'error': 'Unknown dimension'}

        # Count total
        count_query = f"SELECT COUNT(*) as c FROM parts_data WHERE {where_clause}"
        total = conn.execute(count_query, params).fetchone()['c']

        # Get paginated records
        offset = (page - 1) * page_size
        data_query = f"SELECT * FROM parts_data WHERE {where_clause} ORDER BY id LIMIT ? OFFSET ?"
        rows = conn.execute(data_query, params + [page_size, offset]).fetchall()

        # Get file names
        file_ids = set(r['file_id'] for r in rows)
        file_names = {}
        if file_ids:
            placeholders = ','.join('?' * len(file_ids))
            file_rows = conn.execute(
                f'SELECT id, original_filename, sheet_name FROM uploaded_files WHERE id IN ({placeholders})',
                list(file_ids)
            ).fetchall()
            file_names = {r['id']: {'filename': r['original_filename'], 'sheet': r['sheet_name']} for r in file_rows}

        conn.close()

        # 返回所有字段（以当前下钻维度作为关键键高亮显示）
        all_fields = col_names  # 所有统一列

        records = []
        for row in rows:
            data = json.loads(row['data'])
            record = {
                '_record_id': row['id'],
                '_source_file': file_names.get(row['file_id'], {}).get('filename', ''),
                '_source_sheet': file_names.get(row['file_id'], {}).get('sheet', ''),
            }
            for f in all_fields:
                record[f] = data.get(f, '')
            records.append(record)

        return {
            'records': records,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'dimension': dimension,
            'value': value,
            'title': title_desc,
            'all_fields': all_fields,
            'col_display': col_display,
        }

    # ===== 记录对比 =====

    def compare_records(self, field_name, value1, value2):
        """
        对比两条记录的所有字段
        返回: {record1, record2, differences, all_fields}
        """
        conn = get_db()

        # 搜索两条记录 (字段名经 _json_field 校验, 兼容双引擎)
        try:
            field_expr = _json_field(field_name)
        except ValueError:
            conn.close()
            return {
                'success': False,
                'error': '非法字段名',
                'found1': False,
                'found2': False
            }
        query = f"SELECT * FROM parts_data WHERE CAST({field_expr} AS TEXT) LIKE ? LIMIT 1"
        row1 = conn.execute(query, (f'%{value1}%',)).fetchone()
        row2 = conn.execute(query, (f'%{value2}%',)).fetchone()

        if not row1 or not row2:
            conn.close()
            return {
                'success': False,
                'error': '未找到一条或多条记录',
                'found1': row1 is not None,
                'found2': row2 is not None
            }

        data1 = json.loads(row1['data'])
        data2 = json.loads(row2['data'])

        # 添加元信息
        data1['_record_id'] = row1['id']
        data1['_source_file_id'] = row1['file_id']
        data2['_record_id'] = row2['id']
        data2['_source_file_id'] = row2['file_id']

        # 获取文件名
        file_ids = {row1['file_id'], row2['file_id']}
        placeholders = ','.join('?' * len(file_ids))
        file_rows = conn.execute(
            f'SELECT id, original_filename, sheet_name FROM uploaded_files WHERE id IN ({placeholders})',
            list(file_ids)
        ).fetchall()
        file_map = {r['id']: {'filename': r['original_filename'], 'sheet': r['sheet_name']} for r in file_rows}
        conn.close()

        data1['_source_file'] = file_map.get(row1['file_id'], {}).get('filename', '')
        data1['_source_sheet'] = file_map.get(row1['file_id'], {}).get('sheet', '')
        data2['_source_file'] = file_map.get(row2['file_id'], {}).get('filename', '')
        data2['_source_sheet'] = file_map.get(row2['file_id'], {}).get('sheet', '')

        # 收集所有字段（保持顺序）
        all_keys = []
        seen_keys = set()
        for k in list(data1.keys()) + list(data2.keys()):
            if k.startswith('_'):
                continue
            if k not in seen_keys:
                all_keys.append(k)
                seen_keys.add(k)

        # 比较每个字段
        differences = []
        for key in all_keys:
            val1 = data1.get(key, '')
            val2 = data2.get(key, '')
            v1 = str(val1).strip() if val1 else ''
            v2 = str(val2).strip() if val2 else ''
            if v1 != v2:
                differences.append({
                    'field': key,
                    'value1': v1,
                    'value2': v2
                })

        return {
            'success': True,
            'field_name': field_name,
            'value1': value1,
            'value2': value2,
            'record1': data1,
            'record2': data2,
            'differences': differences,
            'all_fields': all_keys,
            'total_fields': len(all_keys),
            'diff_count': len(differences),
            'same_count': len(all_keys) - len(differences)
        }

    def search_complex(self, conditions):
        """
        复杂条件搜索
        conditions: [{field, value, operator}] operator: 'eq'|'neq'|'like'|'not_null'|'is_null'
        """
        conn = get_db()
        where_parts = []
        params = []

        for cond in conditions:
            field = cond.get('field', '')
            value = cond.get('value', '')
            operator = cond.get('operator', 'like')

            # 用户可控字段名: 校验不合法则跳过该条件, 防止标识符注入
            try:
                field_expr = _json_field(field)
            except ValueError:
                print(f"[DB] 忽略非法字段名: {field!r}")
                continue

            if operator == 'eq':
                where_parts.append(f"{field_expr} = ?")
                params.append(value)
            elif operator == 'neq':
                where_parts.append(f"{field_expr} != ?")
                params.append(value)
            elif operator == 'like':
                where_parts.append(f"CAST({field_expr} AS TEXT) LIKE ?")
                params.append(f'%{value}%')
            elif operator == 'not_null':
                where_parts.append(f"{field_expr} IS NOT NULL AND {field_expr} != ''")
            elif operator == 'is_null':
                where_parts.append(f"{field_expr} IS NULL OR {field_expr} = ''")

        where_clause = ' AND '.join(where_parts) if where_parts else '1=1'
        query = f"SELECT * FROM parts_data WHERE {where_clause}"

        rows = conn.execute(query, params).fetchall()

        results = []
        for row in rows:
            data = json.loads(row['data'])
            data['_file_id'] = row['file_id']
            data['_row_number'] = row['row_number']
            data['_record_id'] = row['id']
            results.append(data)

        file_ids = set(r['_file_id'] for r in results)
        file_names = {}
        if file_ids:
            placeholders = ','.join('?' * len(file_ids))
            file_rows = conn.execute(
                f'SELECT id, original_filename, sheet_name FROM uploaded_files WHERE id IN ({placeholders})',
                list(file_ids)
            ).fetchall()
            file_names = {r['id']: {'filename': r['original_filename'], 'sheet': r['sheet_name']} for r in file_rows}

        conn.close()

        for r in results:
            fid = r['_file_id']
            r['_source_file'] = file_names.get(fid, {}).get('filename', '')
            r['_source_sheet'] = file_names.get(fid, {}).get('sheet', '')

        return results

    # ==================== Delta 计算 ====================

    def calculate_delta(self, from_stage="pre-TO", to_stage="TO1",
                         change_filter=None, part_number=None, page=1, page_size=50):
        """计算两个阶段间的 Delta (PN+ZGS组合对比)。

        对比规则：
        - 同PN + 同ZGS = 无delta（两阶段物料完全一样），跳过
        - 同PN + 不同ZGS = ZGS升级delta
        - 后阶段有PN但前阶段没有 = 新增零件delta
        - 前阶段有PN但后阶段没有 = PN停用delta
        - BOM开发顺序: pre-TO → TO1 → TO2

        Delta 检测规则：
        - PN差异：识别后阶段新增/前阶段停用的零件号
        - ZGS变更：同PN比较ZGS值，不同则标记为ZGS升级
        - EC检测：检查后阶段PN在ENIGMA记录中是否存在EC(BuendelNr)
        - ZEUS/FAV：有EC的PN需验证ZEUS(FAV)信息是否已更新
        """
        from config import DELTA_FIELD_CONFIG, DELTA_STAGE_FIELD
        conn = get_db()
        col_names = [r['english_name'] for r in conn.execute(
            'SELECT english_name FROM unified_columns').fetchall()]

        # 1. 确定阶段标识字段存在
        stage_field = DELTA_STAGE_FIELD
        if stage_field not in col_names:
            conn.close()
            return {"success": False, "error": f"Stage field '{stage_field}' not found"}

        # 2-4. 构建两阶段的PN map（复用 _build_stage_pn_map）
        from_map = self._build_stage_pn_map(conn, from_stage, part_number)
        to_map = self._build_stage_pn_map(conn, to_stage, part_number)
        conn.close()

        # 1b. Delta 数据校验：两阶段数据均不能为空
        if not from_map:
            return {"success": False, "error": f"Delta 数据校验失败：前阶段 '{from_stage}' 无BOM数据，请先导入该阶段数据"}
        if not to_map:
            return {"success": False, "error": f"Delta 数据校验失败：后阶段 '{to_stage}' 无BOM数据，请先导入该阶段数据"}

        # 5. PN+ZGS 组合对比（复用 _compute_delta_pairs）
        delta_pairs = self._compute_delta_pairs(from_map, to_map)

        # 构建完整的 delta 对象（含字段级变化详情，用于下钻展示）
        deltas = []
        for pair in delta_pairs:
            delta = self._build_delta(
                pair['part_number'], pair['from_info'], pair['to_info'],
                pair['match_type'], DELTA_FIELD_CONFIG, col_names)
            # EC 检测：检查后阶段 PN 在 ENIGMA 记录中是否存在 EC
            to_data = pair['to_info']['data'] if pair['to_info'] else {}
            ec_value = str(to_data.get('BuendelNr', '')).strip()
            delta['has_ec'] = bool(ec_value)
            delta['ec_value'] = ec_value
            # ZEUS/FAV 信息更新验证：有EC的PN需验证FAV(ZEUS ID)是否已填写
            fav_value = str(to_data.get('FAV_fav', '')).strip()
            delta['has_zeus'] = bool(fav_value)
            delta['zeus_updated'] = bool(ec_value and fav_value)
            delta['fav_value'] = fav_value
            deltas.append(delta)

        # 6. 排序：ZGS升级优先，然后新增，最后停用
        type_order = {"zgs_upgraded": 0, "new_part": 1, "discontinued_part": 2}
        deltas.sort(key=lambda d: (type_order.get(d['match_type'], 9), d['part_number']))

        # 7. 筛选
        if change_filter:
            filter_set = set(change_filter)
            deltas = [d for d in deltas if any(
                c['business'] in filter_set and c['change_type'] not in ('unchanged', 'persisted', 'unavailable')
                for c in d['changes']
            )]

        # 8. 统计摘要
        summary = self._compute_delta_summary(deltas)

        # 9. 分页
        total = len(deltas)
        total_pages = (total + page_size - 1) // page_size
        start = (page - 1) * page_size
        paged = deltas[start:start + page_size]

        return {
            "success": True,
            "deltas": paged,
            "summary": summary,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "from_stage": from_stage,
            "to_stage": to_stage,
        }

    def _build_delta(self, pn, from_info, to_info, match_type,
                      field_config, col_names):
        """构建单个零件的 Delta，包含enigma完整数据用于下钻。"""
        from_data = from_info['data'] if from_info else {}
        to_data = to_info['data'] if to_info else {}

        changes = []
        for cfg in field_config:
            field = cfg['field']
            if field not in col_names and cfg['track']:
                changes.append({
                    "field": field,
                    "business": cfg['business'],
                    "priority": cfg['priority'],
                    "old_value": "",
                    "new_value": "",
                    "change_type": "unavailable"
                })
                continue
            if field not in col_names:
                continue
            old_val = str(from_data.get(field, '')).strip() if from_data else ''
            new_val = str(to_data.get(field, '')).strip() if to_data else ''
            change_type = self._determine_change_type(field, old_val, new_val, cfg['business'])
            changes.append({
                "field": field,
                "business": cfg['business'],
                "priority": cfg['priority'],
                "old_value": old_val,
                "new_value": new_val,
                "change_type": change_type
            })
        changes.sort(key=lambda c: (c['priority'], 0 if c['change_type'] != 'unchanged' else 1))

        # enigma数据（完整记录信息，用于下钻查看详情）
        enigma = {}
        if from_info:
            enigma['from_record'] = from_info['data']
        if to_info:
            enigma['to_record'] = to_info['data']

        return {
            "part_number": pn,
            "match_type": match_type,
            "changes": changes,
            "has_changes": True,
            "enigma": enigma,
            "record_id_old": from_info['id'] if from_info else None,
            "record_id_new": to_info['id'] if to_info else None,
        }

    def _match_stage(self, baulos_val):
        """根据Baulos_aggr字段值匹配阶段名称。"""
        from config import DELTA_STAGE_PATTERNS
        val = str(baulos_val or '').upper()
        for stage_name, pattern in DELTA_STAGE_PATTERNS.items():
            if pattern is None:
                # pre-TO: 既不含PRO1也不含PRO2
                if 'PRO1' not in val and 'PRO2' not in val:
                    return stage_name
            else:
                # SQL LIKE %xxx% 转 Python in 操作
                keyword = pattern.replace('%', '')
                if keyword in val:
                    return stage_name
        # 默认归为pre-TO
        return 'pre-TO'

    def get_delta_detail(self, part_number, from_stage=None, to_stage=None):
        """获取Delta详情（下钻数据）：两阶段并排对比，高亮差异字段。

        参数:
            part_number: 零件号
            from_stage: 前阶段 (pre-TO/TO1/TO2)，可选
            to_stage: 后阶段 (pre-TO/TO1/TO2)，可选

        返回:
            两阶段数据对比，标记差异字段
        """
        from config import DELTA_FIELD_CONFIG, DELTA_STAGE_PATTERNS, DELTA_STAGE_FIELD

        conn = get_db()
        rows = conn.execute(
            "SELECT id, file_id, part_number, data FROM parts_data "
            "WHERE part_number = ? ORDER BY id",
            [part_number]
        ).fetchall()
        all_cols = [r['english_name'] for r in conn.execute(
            'SELECT english_name, display_name FROM unified_columns').fetchall()]
        conn.close()

        if not rows:
            return {"error": "Part not found"}

        # 按阶段分组
        stage_data = {}  # stage_name -> {file_id, data}
        for r in rows:
            data = json.loads(r['data'])
            baulos_val = data.get(DELTA_STAGE_FIELD, '')
            matched_stage = self._match_stage(baulos_val)
            # 同一阶段只保留一条（取最新的）
            stage_data[matched_stage] = {
                'file_id': r['file_id'],
                'data': data,
            }

        # 如果指定了阶段，构建对比
        comparison = None
        if from_stage and to_stage:
            from_data = stage_data.get(from_stage, {}).get('data', {})
            to_data = stage_data.get(to_stage, {}).get('data', {})

            fields = []
            for cfg in DELTA_FIELD_CONFIG:
                field_name = cfg['field']
                # 尝试用不同的大小写匹配字段
                from_val = None
                to_val = None
                for col in all_cols:
                    if col.lower().replace('_', ' ').replace('  ', ' ') == field_name.lower().replace('_', ' '):
                        from_val = from_data.get(col)
                        to_val = to_data.get(col)
                        field_display = col
                        break
                else:
                    # 直接查找
                    from_val = from_data.get(field_name)
                    to_val = to_data.get(field_name)
                    field_display = field_name

                change_type = self._determine_change_type(
                    field_name, from_val, to_val, cfg['business'])

                fields.append({
                    'business': cfg['business'],
                    'field': field_display,
                    'from_value': from_val or '',
                    'to_value': to_val or '',
                    'change_type': change_type,
                    'is_different': change_type not in ['unchanged', 'persisted', 'unavailable'],
                    'priority': cfg['priority'],
                })

            # 额外补充字段（数据库中存在但不在DELTA_FIELD_CONFIG中的）
            extra_fields = []
            all_keys = set(from_data.keys()) | set(to_data.keys())
            config_fields_lower = {cfg['field'].lower().replace('_', ' ') for cfg in DELTA_FIELD_CONFIG}
            for key in sorted(all_keys):
                key_norm = key.lower().replace('_', ' ')
                if key_norm in config_fields_lower:
                    continue
                if key in ['Part Number', 'part_number', 'id']:
                    continue
                from_val = from_data.get(key, '')
                to_val = to_data.get(key, '')
                is_diff = (from_val or '') != (to_val or '')
                if is_diff:  # 只展示有差异的额外字段
                    extra_fields.append({
                        'business': key,
                        'field': key,
                        'from_value': from_val or '',
                        'to_value': to_val or '',
                        'change_type': 'changed' if is_diff else 'unchanged',
                        'is_different': is_diff,
                        'priority': 99,
                    })

            comparison = {
                'from_stage': from_stage,
                'to_stage': to_stage,
                'from_exists': from_stage in stage_data,
                'to_exists': to_stage in stage_data,
                'fields': fields,
                'extra_fields': extra_fields,
                'total_differences': sum(1 for f in fields if f['is_different']) + len(extra_fields),
            }

        return {
            "part_number": part_number,
            "stages_available": list(stage_data.keys()),
            "comparison": comparison,
            "all_columns": all_cols,
            "all_records": [
                {"stage": s, "data": d['data']} for s, d in stage_data.items()
            ],
        }

    def _get_stage_where_clause(self, stage):
        """构建阶段查询的WHERE子句。"""
        from config import DELTA_STAGE_FIELD, DELTA_STAGE_PATTERNS
        stage_field = DELTA_STAGE_FIELD
        pattern = DELTA_STAGE_PATTERNS.get(stage)
        if pattern is None:
            return (f"(UPPER({_json_field(stage_field)}) NOT LIKE '%PRO1%' "
                    f"AND UPPER({_json_field(stage_field)}) NOT LIKE '%PRO2%')")
        else:
            return f"UPPER({_json_field(stage_field)}) LIKE '{pattern}'"

    def _build_stage_pn_map(self, conn, stage, part_number=None):
        """获取某阶段所有PN的map（key=pn, value={zgs, data, id}）。

        复用 calculate_delta 中第3-4步的逻辑。
        同PN多条记录时只保留第一条（按id排序）。

        参数:
            conn: 数据库连接
            stage: 阶段名 (pre-TO/TO1/TO2)
            part_number: 可选的PN过滤（模糊匹配）

        返回:
            {pn: {id, zgs, data}} 的字典
        """
        clause = self._get_stage_where_clause(stage)
        pn_filter = ""
        pn_params = []
        if part_number:
            pn_filter = " AND part_number LIKE ?"
            pn_params = [f'%{part_number}%']

        rows = conn.execute(
            f"SELECT id, part_number, data FROM parts_data "
            f"WHERE {clause}{pn_filter} ORDER BY id", pn_params
        ).fetchall()

        pn_map = {}
        for r in rows:
            pn = r['part_number']
            if pn and pn not in pn_map:
                data = json.loads(r['data'])
                zgs = str(data.get('ZGS DiaP', '')).strip()
                pn_map[pn] = {'id': r['id'], 'zgs': zgs, 'data': data}
        return pn_map

    def _compute_delta_pairs(self, from_map, to_map):
        """PN+ZGS 组合对比，返回 delta pairs 列表。

        复用 calculate_delta 中第5步的对比逻辑：
        - 同PN + 同ZGS = 无delta（跳过）
        - 同PN + 不同ZGS = ZGS升级delta (zgs_upgraded)
        - 后阶段有PN前阶段没有 = 新增零件delta (new_part)
        - 前阶段有PN后阶段没有 = PN停用delta (discontinued_part)

        参数:
            from_map: 前阶段PN map {pn: {id, zgs, data}}
            to_map: 后阶段PN map {pn: {id, zgs, data}}

        返回:
            deltas 列表，每个元素包含:
            {part_number, match_type, from_info, to_info}
            match_type: "zgs_upgraded" | "new_part" | "discontinued_part"
        """
        deltas = []

        # 5a. 同PN比较ZGS: 同PN+同ZGS=跳过, 同PN+不同ZGS=升级delta
        for pn in sorted(set(from_map.keys()) & set(to_map.keys())):
            from_zgs = from_map[pn]['zgs']
            to_zgs = to_map[pn]['zgs']
            if from_zgs == to_zgs:
                continue  # 同PN+同ZGS = 无delta
            deltas.append({
                'part_number': pn,
                'match_type': 'zgs_upgraded',
                'from_info': from_map[pn],
                'to_info': to_map[pn],
            })

        # 5b. 新增零件（仅存在于后阶段）
        for pn in sorted(set(to_map.keys()) - set(from_map.keys())):
            deltas.append({
                'part_number': pn,
                'match_type': 'new_part',
                'from_info': None,
                'to_info': to_map[pn],
            })

        # 5c. PN停用（仅存在于前阶段）
        for pn in sorted(set(from_map.keys()) - set(to_map.keys())):
            deltas.append({
                'part_number': pn,
                'match_type': 'discontinued_part',
                'from_info': from_map[pn],
                'to_info': None,
            })

        return deltas

    def _compute_kpi_from_delta_pairs(self, delta_pairs):
        """从 delta pairs 计算 Dashboard KPI 数据。

        基于PN+ZGS组合匹配的结果，计算各项KPI指标。
        EC检测规则：检查后阶段PN在ENIGMA记录中是否存在EC(BuendelNr)，
        存在即为EC变更（不仅限于EC从无到有的转换）。
        ZEUS/FAV验证：有EC的PN需验证FAV(ZEUS ID)信息是否已更新。

        参数:
            delta_pairs: _compute_delta_pairs 返回的列表

        返回:
            {new_pn, discontinued_pn, zgs_changed, total_delta, new_ec, new_kem, soma_ja, ec_with_zeus}
        """
        new_pn = 0
        discontinued_pn = 0
        zgs_changed = 0
        new_ec = 0
        new_kem = 0
        soma_ja = 0
        ec_with_zeus = 0

        for pair in delta_pairs:
            match_type = pair['match_type']
            from_info = pair['from_info']
            to_info = pair['to_info']

            if match_type == 'new_part':
                new_pn += 1
            elif match_type == 'discontinued_part':
                discontinued_pn += 1
            elif match_type == 'zgs_upgraded':
                zgs_changed += 1

            from_data = from_info['data'] if from_info else {}
            to_data = to_info['data'] if to_info else {}

            # EC检测：后阶段PN在ENIGMA记录中存在EC(BuendelNr)即为EC变更
            to_ec = str(to_data.get('BuendelNr', '')).strip() if to_data else ''
            from_ec = str(from_data.get('BuendelNr', '')).strip() if from_data else ''
            has_ec = bool(to_ec)
            ec_added = not from_ec and to_ec
            if has_ec:
                new_ec += 1

            # KEM从无到有（基于EC新增的KEM释放，KEM是EC的子集）
            from_kem = str(from_data.get('KEM Number', '')).strip() if from_data else ''
            to_kem = str(to_data.get('KEM Number', '')).strip() if to_data else ''
            if ec_added and not from_kem and to_kem:
                new_kem += 1

            # ZEUS/FAV信息更新验证：有EC的PN需验证FAV(ZEUS ID)是否已填写
            to_fav = str(to_data.get('FAV_fav', '')).strip() if to_data else ''
            if has_ec and to_fav:
                ec_with_zeus += 1

            # SOMA in ZEUS 从无到有 / 从'nein'变为'ja'
            from_soma = str(from_data.get('SOMA in ZEUS', '')).strip().lower() if from_data else ''
            to_soma = str(to_data.get('SOMA in ZEUS', '')).strip().lower() if to_data else ''
            if from_soma != 'ja' and to_soma == 'ja':
                soma_ja += 1

        return {
            'new_pn': new_pn,
            'discontinued_pn': discontinued_pn,
            'zgs_changed': zgs_changed,
            'total_delta': len(delta_pairs),
            'new_ec': new_ec,
            'new_kem': new_kem,
            'soma_ja': soma_ja,
            'ec_with_zeus': ec_with_zeus,
        }

    def get_delta_dashboard_data(self):
        """获取Delta可视化面板所需的全部数据。

        返回结构:
            - stages: 各阶段基础统计
            - delta1 / delta2: 各Delta区间的KPI和饼图数据
            - bar_line: 柱状折线图数据
        """
        from config import DELTA_STAGE_PATTERNS, DELTA_STAGE_FIELD
        conn = get_db()

        stage_field = DELTA_STAGE_FIELD
        stages = ['pre-TO', 'TO1', 'TO2']

        # 字段名（与数据库列对应）
        ec_col = 'BuendelNr'
        fav_col = 'FAV_fav'
        fav_status_col = 'FAVStatusKurz_fav'
        soma_col = 'SOMA in ZEUS'
        kem_col = 'KEM Number'
        zgs_col = 'ZGS DiaP'
        ec_status_col = 'Process Status'

        # === 1. 各阶段基础统计 ===
        stage_stats = {}
        for stage in stages:
            clause = self._get_stage_where_clause(stage)
            total = conn.execute(f"SELECT COUNT(*) as c FROM parts_data WHERE {clause}").fetchone()['c']
            pn_count = conn.execute(
                f"SELECT COUNT(DISTINCT part_number) as c FROM parts_data "
                f"WHERE part_number != '' AND {clause}"
            ).fetchone()['c']

            # EC统计
            ec_count = conn.execute(
                f"SELECT COUNT(DISTINCT {_json_field(ec_col)}) as c FROM parts_data "
                f"WHERE {_json_field(ec_col)} IS NOT NULL "
                f"AND {_json_field(ec_col)} != '' AND {clause}"
            ).fetchone()['c']
            ec_pn = conn.execute(
                f"SELECT COUNT(DISTINCT part_number) as c FROM parts_data "
                f"WHERE part_number != '' AND {_json_field(ec_col)} IS NOT NULL "
                f"AND {_json_field(ec_col)} != '' AND {clause}"
            ).fetchone()['c']

            # FAV/ZEUS统计
            fav_count = conn.execute(
                f"SELECT COUNT(DISTINCT {_json_field(fav_col)}) as c FROM parts_data "
                f"WHERE {_json_field(fav_col)} IS NOT NULL "
                f"AND {_json_field(fav_col)} != '' AND {clause}"
            ).fetchone()['c']
            fav_pn = conn.execute(
                f"SELECT COUNT(DISTINCT part_number) as c FROM parts_data "
                f"WHERE part_number != '' AND {_json_field(fav_col)} IS NOT NULL "
                f"AND {_json_field(fav_col)} != '' AND {clause}"
            ).fetchone()['c']

            # KEM统计
            kem_count = conn.execute(
                f"SELECT COUNT(DISTINCT {_json_field(kem_col)}) as c FROM parts_data "
                f"WHERE {_json_field(kem_col)} IS NOT NULL "
                f"AND {_json_field(kem_col)} != '' AND {clause}"
            ).fetchone()['c']

            # SOMA=Ja统计
            soma_ja = conn.execute(
                f"SELECT COUNT(*) as c FROM parts_data "
                f"WHERE LOWER({_json_field(soma_col)}) = 'ja' AND {clause}"
            ).fetchone()['c']

            stage_stats[stage] = {
                'total_records': total,
                'unique_pn': pn_count,
                'ec_count': ec_count,
                'ec_pn': ec_pn,
                'fav_count': fav_count,
                'fav_pn': fav_pn,
                'kem_count': kem_count,
                'soma_ja': soma_ja,
            }

        # === 2. Delta KPI 计算（两阶段对比，基于PN+ZGS组合匹配算法） ===
        # 使用与 calculate_delta 完全一致的 PN+ZGS 组合匹配逻辑
        pre_to_map = self._build_stage_pn_map(conn, 'pre-TO')
        to1_map = self._build_stage_pn_map(conn, 'TO1')
        to2_map = self._build_stage_pn_map(conn, 'TO2')

        delta1_pairs = self._compute_delta_pairs(pre_to_map, to1_map)
        delta2_pairs = self._compute_delta_pairs(to1_map, to2_map)

        delta1_kpi_full = self._compute_kpi_from_delta_pairs(delta1_pairs)
        delta2_kpi_full = self._compute_kpi_from_delta_pairs(delta2_pairs)

        # 保持返回字段与前端一致（去掉total_delta，与原有结构对齐）
        delta1_kpi = {k: v for k, v in delta1_kpi_full.items() if k != 'total_delta'}
        delta2_kpi = {k: v for k, v in delta2_kpi_full.items() if k != 'total_delta'}

        # Delta 数据校验：确保两阶段数据均非空
        valid = bool(pre_to_map) and bool(to1_map) and bool(to2_map)

        # === 3. EC Process Status 饼图数据 ===
        def get_ec_status_distribution(stage):
            clause = self._get_stage_where_clause(stage)
            rows = conn.execute(
                f"SELECT {_json_field(ec_status_col)} as val, COUNT(*) as c "
                f"FROM parts_data "
                f"WHERE {_json_field(ec_status_col)} IS NOT NULL "
                f"AND {_json_field(ec_status_col)} != '' "
                f"AND {clause} GROUP BY {_json_field(ec_status_col)} "
                f"ORDER BY c DESC"
            ).fetchall()
            return [{'name': r['val'], 'value': r['c']} for r in rows if r['val']]

        ec_pie_to1 = get_ec_status_distribution('TO1')
        ec_pie_to2 = get_ec_status_distribution('TO2')

        # === 4. FAV Status 饼图数据 ===
        def get_fav_status_distribution(stage):
            clause = self._get_stage_where_clause(stage)
            rows = conn.execute(
                f"SELECT {_json_field(fav_status_col)} as val, COUNT(*) as c "
                f"FROM parts_data "
                f"WHERE {_json_field(fav_status_col)} IS NOT NULL "
                f"AND {_json_field(fav_status_col)} != '' "
                f"AND {clause} GROUP BY {_json_field(fav_status_col)} "
                f"ORDER BY c DESC"
            ).fetchall()
            return [{'name': r['val'], 'value': r['c']} for r in rows if r['val']]

        fav_pie_to1 = get_fav_status_distribution('TO1')
        fav_pie_to2 = get_fav_status_distribution('TO2')

        # === 5. 柱状折线图数据 ===
        bar_line = {
            'stages': stages,
            'ec_counts': [stage_stats[s]['ec_pn'] for s in stages],
            'fav_counts': [stage_stats[s]['fav_pn'] for s in stages],
        }

        conn.close()

        return {
            'valid': valid,
            'stages': stage_stats,
            'delta1': {
                'label': 'pre-TO → TO1',
                'kpi': delta1_kpi,
                'ec_pie': ec_pie_to1,
                'fav_pie': fav_pie_to1,
            },
            'delta2': {
                'label': 'TO1 → TO2',
                'kpi': delta2_kpi,
                'ec_pie': ec_pie_to2,
                'fav_pie': fav_pie_to2,
            },
            'bar_line': bar_line,
        }

    def _determine_change_type(self, field, old_val, new_val, business_name):
        """判断字段变化类型。"""
        # ZGS 升级检测
        if business_name == "ZGS" and old_val and new_val:
            try:
                if int(new_val) > int(old_val):
                    return "upgraded"
                elif int(new_val) < int(old_val):
                    return "changed"
                else:
                    return "persisted"
            except ValueError:
                pass
        # 通用判断
        if not old_val and new_val:
            return "added"
        if old_val and not new_val:
            return "removed"
        if old_val and new_val and old_val != new_val:
            return "changed"
        if old_val and new_val and old_val == new_val:
            return "persisted"
        return "unchanged"

    def _compute_delta_summary(self, deltas):
        """计算 Delta 统计摘要。"""
        summary = {
            "total_records": len(deltas),
            "zgs_upgraded": 0,
            "ec_added": 0,
            "new_parts": 0,
            "discontinued_parts": 0,
            "has_ec": 0,
            "zeus_updated": 0,
        }
        for d in deltas:
            if d['match_type'] == 'zgs_upgraded':
                summary['zgs_upgraded'] += 1
            elif d['match_type'] == 'new_part':
                summary['new_parts'] += 1
            elif d['match_type'] == 'discontinued_part':
                summary['discontinued_parts'] += 1
            for c in d['changes']:
                if c['business'] == 'EC' and c['change_type'] == 'added':
                    summary['ec_added'] += 1
            if d.get('has_ec'):
                summary['has_ec'] += 1
            if d.get('zeus_updated'):
                summary['zeus_updated'] += 1
        return summary


# 全局实例
db_manager = DatabaseManager()
